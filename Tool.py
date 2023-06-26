#!/usr/bin/python
# -*- coding: UTF-8 -*-

import os
from openpyxl import Workbook
from openpyxl import load_workbook
import hashlib
import base64
import arrow
import re

class ExcelTool(object):
    __file_name = "timeTask.xlsx"
    __sheet_name = "定时任务"
    __dir_name = "taskFile"
    
    # 新建工作簿
    def create_excel(self, file_name: str = __file_name, sheet_name=__sheet_name):
        # 文件路径
        workbook_file_path = self.get_file_path(file_name)

        # 创建Excel
        if not os.path.exists(workbook_file_path):
            wb = Workbook()
            wb.create_sheet(sheet_name, 0)
            wb.save(workbook_file_path)
            print("定时Excel创建成功，文件路径为：{}".format(workbook_file_path))
        else:
            print("timeTask文件已存在, 无需创建")
                

    # 读取内容,返回元组列表
    def readExcel(self, file_name=__file_name, sheet_name=__sheet_name):
        # 文件路径
        workbook_file_path = self.get_file_path(file_name)
        
        # 文件存在
        if os.path.exists(workbook_file_path):
            wb = load_workbook(workbook_file_path)
            ws = wb[sheet_name]
            data = list(ws.values)
            #print(data)
            return data
        else:
            print("timeTask文件不存在, 读取数据为空")
            return []


    # 写入列表，返回元组列表
    def addItemToExcel(self, item, file_name=__file_name, sheet_name=__sheet_name):
        # 文件路径
        workbook_file_path = self.get_file_path(file_name)
        
        # 如果文件存在,就执行
        if os.path.exists(workbook_file_path):
            wb = load_workbook(workbook_file_path)
            ws = wb[sheet_name]
            ws.append(item)
            wb.save(workbook_file_path)
            
            # 列表
            data = list(ws.values)
            #print(data)
            return data
        else:
            print("timeTask文件不存在, 添加数据失败")
            return []
        
        
    # 置为失效
    def disableItemToExcel(self, taskId, file_name=__file_name, sheet_name=__sheet_name):
        #读取数据
        data = self.readExcel(file_name, sheet_name)
        if len(data) > 0:
            # 表格对象
            workbook_file_path = self.get_file_path(file_name)
            wb = load_workbook(workbook_file_path)
            ws = wb[sheet_name]
            isExist = False
            taskContent = ""
            #遍历
            for index, hisItem in enumerate(data):
                 #ID是否相同
                 if hisItem[0] == taskId:
                    #置为已消费：即0
                    ws.cell(index + 1, 2).value = "0"
                    isExist = True
                    #循环信息 + 时间 + 事件内容
                    taskContent = hisItem[3] + " " + hisItem[2] + " " + hisItem[4]
                    
            if isExist: 
                #保存
                wb.save(workbook_file_path)
            
            return isExist, taskContent
        else:
            print("timeTask文件无数据, 消费数据失败")
            return False, ""
    
    
    #获取文件路径      
    def get_file_path(self, file_name=__file_name):
        # 文件路径
        current_file = os.path.abspath(__file__)
        current_dir = os.path.dirname(current_file)
        workbook_file_path = current_dir + "/" + self.__dir_name + "/" + file_name
        
        # 工作簿当前目录
        workbook_dir_path = os.path.dirname(workbook_file_path)
        # 创建目录
        if not os.path.exists(workbook_dir_path):
            # 创建工作簿路径,makedirs可以创建级联路径
            os.makedirs(workbook_dir_path)
            
        return workbook_file_path
        

#task模型        
class TimeTaskModel:
    #Item数据排序
    #0：ID - 唯一ID (自动生成，无需填写)
    #1：是否可用 - 0/1，0=不可用，1=可用
    #2：时间信息 - 格式为：HH:mm:ss
    #3：轮询信息 - 格式为：每天、每周N、YYYY-MM-DD
    #4：消息内容 - 消息内容
    #5：fromUser - 来源user
    #6：toUser - 发送给的user
    #7：other_user_id - otehrID
    #8：isGroup - 0/1，是否群聊； 0=否，1=是
    #9：原始内容 - 原始的消息体
    
    def __init__(self, item, isNeedFormat: bool):
        
        self.taskId = item[0]
        self.enable = item[1] == "1"
        self.timeStr = item[2]
        self.circleTimeStr = item[3]
        self.eventStr = item[4]
        self.fromUser = item[5]
        self.toUser = item[6]
        self.other_user_id = item[7]
        self.isGroup = item[8] == "1"
        self.originMsg = item[9]
        
        #需要处理格式
        if isNeedFormat:
            #计算内容ID (使用不可变的内容计算，去除元素：enable 会变、originMsg中有时间戳)
            new_tuple = (self.timeStr, self.circleTimeStr, self.eventStr, self.fromUser, 
                         self.toUser, self.other_user_id, "1" if self.isGroup else "0")
            temp_content='_'.join(new_tuple)
            short_id = self.get_short_id(temp_content)
            print(f'消息体：{temp_content}， 唯一ID：{short_id}')
            self.taskId = short_id
            
            #入库的周期、time
            g_circle = self.get_cicleDay(self.circleTimeStr)
            g_time = self.get_time(self.timeStr)
            self.timeStr = g_time
            self.circleTimeStr = g_circle
        
        
    #获取格式化后的Item
    def get_formatItem(self):
        temp_item = (self.taskId,
                "1" if self.enable else "0",
                self.timeStr,
                self.circleTimeStr,
                self.eventStr,
                self.fromUser,
                self.toUser,
                self.other_user_id,
                "1" if self.isGroup else "0",
                self.originMsg) 
        return temp_item
            
    #计算唯一ID        
    def get_short_id(self, string):
        # 使用 MD5 哈希算法计算字符串的哈希值
        hash_value = hashlib.md5(string.encode()).digest()
    
        # 将哈希值转换为一个 64 进制的短字符串
        short_id = base64.urlsafe_b64encode(hash_value)[:8].decode()
        return short_id
    
    
    #是否现在时间      
    def is_nowTime(self):
        tempTimeStr = self.timeStr
        #如果以00结尾，对比精准度为分钟
        if tempTimeStr.count(":") == 2 and tempTimeStr.endswith("00"):
           return (arrow.now().format('HH:mm') + ":00") == tempTimeStr
        #对比精准到秒 
        tempValue = arrow.now().format('HH:mm:ss') == tempTimeStr
        return tempValue 
    
    #是否未来时间      
    def is_featureTime(self):
        tempTimeStr = self.timeStr
        tempValue = arrow.get(tempTimeStr, 'HH:mm:ss').time() > arrow.now().time()
        return tempValue 
    
    #是否未来day      
    def is_featureDay(self):
        tempStr = self.circleTimeStr
        tempValue = "每周" in tempStr or "每星期" in tempStr or "每天" in tempStr  or "工作日" in tempStr
        #日期
        if self.is_valid_date(tempStr):
            tempValue = arrow.get(tempStr, 'YYYY-MM-DD').datetime > arrow.now().datetime
            
        return tempValue 
    
    #是否today      
    def is_today(self):
        #当前时间
        current_time = arrow.now()
        #轮询信息
        item_circle = self.circleTimeStr
        if self.is_valid_date(item_circle):
            #日期相等
            if item_circle == current_time.format('YYYY-MM-DD'):
                #今天要出发的任务
                print(f"[定时任务]类型: 录入日期, 日期信息：{item_circle}")
                return True
            else:
                #其他时间待出发
                print(f"[定时任务]类型: 录入日期, 非今天任务, 日期信息：{item_circle}")
                return False
            
        elif "每天" in item_circle:
            #今天要出发的任务
            print(f"[定时任务]类型：每天")
            return True
        
        elif "每周" in item_circle or "每星期" in item_circle:
            if self.is_today_weekday(item_circle):
                print(f"[定时任务]类型: 每周, 日期信息：{item_circle}")
                return True
            else:
                print(f"[定时任务]类型: 每周, 非今天任务, 日期信息为：{item_circle}")
                return False    
            
        elif "工作日" in item_circle:
                # 判断星期几
                weekday = arrow.now().weekday()
                # 判断是否是工作日
                is_weekday = weekday < 5
                if is_weekday:
                    print(f"[定时任务]类型: 工作日")
                    return True
                else:
                    print(f"[定时任务]类型: 工作日, 非今天任务，日期信息为：{item_circle}")
                    return False    
                    
    #是否今天的星期数       
    def is_today_weekday(self, weekday_str):
        # 将中文数字转换为阿拉伯数字
        weekday_dict = {'一': 1, '二': 2, '三': 3, '四': 4, '五': 5, '六': 6, '日': 7}
        weekday_num = weekday_dict.get(weekday_str[-1])
        if weekday_num is None:
            return False
        
        # 判断今天是否是指定的星期几
        today = arrow.now()
        tempValue = today.weekday() == weekday_num - 1   
        return tempValue   
        
    #日期是否格式正确
    def is_valid_date(self, date_string):
        pattern = re.compile(r'^\d{4}-\d{2}-\d{2}$')
        match = pattern.match(date_string)
        return match is not None
    
    
    #获取周期
    def get_cicleDay(self, circleStr):
        # 定义正则表达式
        pattern = r'^\d{4}-\d{2}-\d{2}$'
        # 是否符合 YYYY-MM-DD 格式的日期
        isGoodDay = re.match(pattern, circleStr)
        
        g_circle = ""
        #如果可被解析为具体日期
        if circleStr in ['今天', '明天', '后天']:
              #今天
              today = arrow.now('local')
              if circleStr == '今天':
                    # 将日期格式化为 YYYY-MM-DD 的格式
                    formatted_today = today.format('YYYY-MM-DD')
                    g_circle = formatted_today
                    
              elif circleStr == '明天':
                    tomorrow = today.shift(days=1)
                    formatted_tomorrow = tomorrow.format('YYYY-MM-DD')
                    g_circle = formatted_tomorrow
                    
              elif circleStr == '后天':
                    after_tomorrow = today.shift(days=2)
                    formatted_after_tomorrow = after_tomorrow.format('YYYY-MM-DD')
                    g_circle = formatted_after_tomorrow
              else:
                  print('暂不支持的格式')
                   
                    
        #YYYY-MM-DD 格式
        elif isGoodDay:
            g_circle = circleStr
            
        #每天、每周、工作日
        elif circleStr in ["每天", "每周", "工作日"]:
                g_circle = circleStr
        
        #每周X
        elif circleStr in ["每周一", "每周二", "每周三", "每周四", "每周五", "每周六","每周日","每周天", 
                           "每星期一", "每星期二","每星期三", "每星期四", "每星期五","每星期六", "每星期日", "每星期天"]:       
            #每天、每周X等
            g_circle = circleStr
            
        else:
            print('暂不支持的格式')
            
        return g_circle
    
    #获取时间
    def get_time(self, timeStr):
        pattern1 = r'^\d{2}:\d{2}:\d{2}$'
        pattern2 = r'^\d{2}:\d{2}$'
        # 是否符合 HH:mm:ss 格式
        time_good1 = re.match(pattern1, timeStr)
        # 是否符合 HH:mm 格式
        time_good2 = re.match(pattern2, timeStr)
        
        g_time = ""
        if time_good1 :
            g_time = timeStr
            
        elif time_good2:
            g_time = timeStr + ":00"
        
        elif '点' in timeStr or '分' in timeStr or '秒' in timeStr :
            content = timeStr.replace("点", ":")
            content = content.replace("分", ":")
            content = content.replace("秒", "")
            wordsArray = content.split(":")
            hour = "0"
            minute = "0"
            second = "0"
            digits = {'零': 0, '一': 1, '二': 2, '三': 3, '四': 4, '五': 5, '六': 6, '七': 7, '八': 8, '九': 9, '十': 10, 
                '十一': 11, '十二': 12, '十三': 13, '十四': 14, '十五': 15, '十六': 16, '十七': 17, '十八': 18, '十九': 19, '二十': 20, 
                '二十一': 21, '二十二': 22, '二十三': 23, '二十四': 24, '二十五': 25, '二十六': 26, '二十七': 27, '二十八': 28, '二十九': 29, '三十': 30, 
                '三十一': 31, '三十二': 32, '三十三': 33, '三十四': 34, '三十五': 35, '三十六': 36, '三十七': 37, '三十八': 38, '三十九': 39, '四十': 40, 
                '四十一': 41, '四十二': 42, '四十三': 43, '四十四': 44, '四十五': 45, '四十六': 46, '四十七': 47, '四十八': 48, '四十九': 49, '五十': 50, 
                '五十一': 51, '五十二': 52, '五十三': 53, '五十四': 54, '五十五': 55, '五十六': 56, '五十七': 57, '五十八': 58, '五十九': 59, '六十': 60, '半': 30}
            for index, item in enumerate(wordsArray):
                if index == 0 and len(item) > 0:
                    #中文 且 在一 至 六十之间
                    if re.search('[\u4e00-\u9fa5]', item) and item in digits.keys():
                        hour = str(digits[item])
                    elif item in digits.values():
                         hour = str(item)
                    else:
                        return ""       
                            
                elif index == 1 and len(item) > 0:
                    if re.search('[\u4e00-\u9fa5]', item) and item in digits.keys():
                        minute = str(digits[item])
                    elif item in digits.values():
                        minute = str(item)
                    else:
                        return ""  
                        
                elif index == 2 and len(item) > 0:
                    if re.search('[\u4e00-\u9fa5]', item) and item in digits.keys():
                        second = str(digits[item])
                    elif item in digits.values():
                        second = str(item)  
                    else:
                        return ""    
            
            #格式处理       
            if int(hour) < 10:
                  hour = "0" + hour
                      
            if int(minute) < 10:
                  minute = "0" + minute
                  
            if int(second) < 10:
                  second = "0" + second  
            
            #拼接     
            g_time = hour + ":" + minute + ":" + second                                       
            
        else:
            print('暂不支持的格式')
            return ""
            
        #检测转换的时间是否合法    
        time_good1 = re.match(pattern1, g_time)
        if time_good1:
              return g_time
                 
        return ""
